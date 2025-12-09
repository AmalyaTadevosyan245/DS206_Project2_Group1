"""
tasks.py
Task-level ETL functions for DS206 Project 2.
Each task:
  - loads SQL template
  - substitutes T-SQL parameters (@param)
  - executes batches (GO separated)
  - returns {"success": True/False, "message": "..."}
"""

import os
import sys
import openpyxl

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from utils import (
    load_db_config,
    create_db_connection,
    read_sql_file,
    render_sql,
    execute_sql
)


# ==============================================================
# Helper — run a single parametrized SQL script
# ==============================================================

def run_sql_task(sql_path: str, params: dict) -> dict:
    """
    Execute a parametrized SQL script.
    Handles:
      - @parameter substitution
      - splitting batches by GO
      - atomicity
    """
    try:
        cfg = load_db_config()
        conn = create_db_connection(cfg)

        sql_text = read_sql_file(sql_path)
        sql_rendered = render_sql(sql_text, params)

        # split into batches
        batches = [
            batch.strip() for batch in sql_rendered.split("\nGO")
            if batch.strip()
        ]

        for batch in batches:
            execute_sql(conn, batch)

        conn.close()
        return {"success": True}

    except Exception as e:
        return {"success": False, "message": str(e)}


# ==============================================================
# 1. Create Dimensional Tables
# ==============================================================

def task_create_dimensional_tables() -> dict:
    sql_path = os.path.join(
        PROJECT_ROOT, "infrastructure_initiation", "dimensional_db_table_creation.sql"
    )

    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo"
    }

    return run_sql_task(sql_path, params)


# ==============================================================
# Dimensional Load Tasks (SCD1 / SCD2 / SCD3 / SCD4)
# ==============================================================

def task_update_dim_categories(prereq=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(PROJECT_ROOT, "pipeline_dimensional_data/queries/update_dim_categories.sql")
    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "dim_table_name": "DimCategories",
        "staging_table_name": "staging_Categories"
    }
    return run_sql_task(sql_path, params)


def task_update_dim_customers(prereq=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(PROJECT_ROOT, "pipeline_dimensional_data/queries/update_dim_customers.sql")
    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "dim_table_name": "DimCustomers",
        "staging_table_name": "staging_Customers"
    }
    return run_sql_task(sql_path, params)


def task_update_dim_employees(prereq=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(PROJECT_ROOT, "pipeline_dimensional_data/queries/update_dim_employees.sql")
    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "dim_table_name": "DimEmployees",
        "staging_table_name": "staging_Employees"
    }
    return run_sql_task(sql_path, params)


def task_update_dim_products(prereq=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(PROJECT_ROOT, "pipeline_dimensional_data/queries/update_dim_products.sql")
    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "dim_table_name": "DimProducts",
        "staging_table_name": "staging_Products"
    }
    return run_sql_task(sql_path, params)


def task_update_dim_region(prereq=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(PROJECT_ROOT, "pipeline_dimensional_data/queries/update_dim_region.sql")
    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "dim_table_name": "DimRegion",
        "staging_table_name": "staging_Region"
    }
    return run_sql_task(sql_path, params)


def task_update_dim_shippers(prereq=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(PROJECT_ROOT, "pipeline_dimensional_data/queries/update_dim_shippers.sql")
    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "dim_table_name": "DimShippers",
        "staging_table_name": "staging_Shippers"
    }
    return run_sql_task(sql_path, params)


def task_update_dim_suppliers(prereq=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(PROJECT_ROOT, "pipeline_dimensional_data/queries/update_dim_suppliers.sql")
    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "dim_table_name": "DimSuppliers",
        "staging_table_name": "staging_Suppliers"
    }
    return run_sql_task(sql_path, params)


def task_update_dim_territories(prereq=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(PROJECT_ROOT, "pipeline_dimensional_data/queries/update_dim_territories.sql")
    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "dim_table_name": "DimTerritories",
        "staging_table_name": "staging_Territories"
    }
    return run_sql_task(sql_path, params)


# ==============================================================
# Fact Tasks (Snapshot Fact + Fact Error)
# ==============================================================

def task_update_factorders(prereq=None, start_date=None, end_date=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(
        PROJECT_ROOT,
        "pipeline_dimensional_data/queries/update_factorders.sql"
    )

    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "fact_table_name": "FactOrders",
        "orders_staging_table": "staging_Orders",
        "details_staging_table": "staging_OrderDetails"
    }

    return run_sql_task(sql_path, params)


def task_update_fact(start_date, end_date, prereq=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(PROJECT_ROOT, "pipeline_dimensional_data/queries/update_fact.sql")
    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "fact_table_name": "FactOrders",
        "staging_table_name": "staging_Orders",
        "start_date": start_date,
        "end_date": end_date
    }
    return run_sql_task(sql_path, params)


def task_update_fact_error(start_date, end_date, prereq=None) -> dict:
    if prereq and not prereq.get("success"):
        return {"success": False, "message": "Prerequisite failed"}

    sql_path = os.path.join(PROJECT_ROOT, "pipeline_dimensional_data/queries/update_fact_error.sql")
    params = {
        "database_name": "ORDER_DDS",
        "schema_name": "dbo",
        "fact_error_table": "FactOrders_Error",
        "staging_table_name": "staging_Orders",
        "start_date": start_date,
        "end_date": end_date
    }

# ==============================================================
# Table Population
# ==============================================================

SOURCE_PATH = "raw_data_source.xlsx"
STAGING_TABLES = {
    "staging_Categories": (SOURCE_PATH, "Categories"),
    "staging_Customers": (SOURCE_PATH, "Customers"),
    "staging_Employees": (SOURCE_PATH, "Employees"),
    "staging_OrderDetails": (SOURCE_PATH, "OrderDetails"),
    "staging_Orders": (SOURCE_PATH, "Orders"),
    "staging_Products": (SOURCE_PATH, "Products"),
    "staging_Region": (SOURCE_PATH, "Region"),
    "staging_Shippers": (SOURCE_PATH, "Shippers"),
    "staging_Suppliers": (SOURCE_PATH, "Suppliers"),
    "staging_Territories": (SOURCE_PATH, "Territories"),
}

def populate_table_from_excel(cursor, table_name, sheet):
    """Insert all rows from an Excel sheet into the corresponding staging table."""
    num_cols = sheet.max_column
    placeholders = ", ".join(["?"] * num_cols)
    query = f"INSERT INTO {table_name} VALUES ({placeholders})"

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        cursor.execute(query, row)
    cursor.commit()


def task_populate_staging(prereq=None) -> dict:
    """
    Populate all staging tables from Excel file.

    Args:
        prereq (dict): prerequisite task result

    Returns:
        dict: {'success': True/False, 'message': "..."}
    """
    if prereq and not prereq.get("success", False):
        return {"success": False, "message": "Prerequisite failed"}

    try:
        cfg = load_db_config()
        conn = create_db_connection(cfg)
        cursor = conn.cursor()

        for table_name, (default_file, sheet_name) in STAGING_TABLES.items():
            file_path = default_file
            if not os.path.exists(file_path):
                print(f"Excel file not found: {file_path}, skipping {table_name}")
                continue

            print(f"Populating {table_name} from {file_path} / sheet {sheet_name}")
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb[sheet_name]
            populate_table_from_excel(cursor, table_name, sheet)

        conn.close()
        return {"success": True}

    except Exception as e:
        return {"success": False, "message": str(e)}
    return run_sql_task(sql_path, params)
